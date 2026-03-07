"""
Import Drawing Metadata from Overview Excel

This script reads the overview document list and extracts:
- Actual PDF filenames from HYPERLINK formulas
- Relative paths from the HYPERLINK
- Document descriptions

Creates a lookup dictionary keyed by filename for use during PDF indexing.
"""

import openpyxl
import re
import os
import sys
import pickle
from database import ShipCableDB


def import_drawing_metadata(excel_path: str) -> dict:
    """
    Import drawing metadata from the overview Excel file.
    Extracts filenames from HYPERLINK formulas in column C.
    
    Returns a dictionary keyed by filename (lowercase for case-insensitive matching).
    """
    print(f"Reading Excel file: {excel_path}")
    
    # Load with data_only=False to see formulas
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    sheet = wb.active
    
    # Pattern to extract path from HYPERLINK formula
    # =HYPERLINK("path\to\file.pdf","display_name") or ==HYPERLINK(...)
    hyperlink_pattern = re.compile(r'=+HYPERLINK\("([^"]+)"')
    
    metadata = {}
    rows_processed = 0
    hyperlinks_found = 0
    
    # Find the header row and data start
    # Headers are in row 5, data starts at row 6
    for row in range(6, sheet.max_row + 1):
        rows_processed += 1
        
        cell = sheet.cell(row=row, column=3)  # Column C = Filename
        formula = cell.value
        
        if formula and 'HYPERLINK' in str(formula):
            match = hyperlink_pattern.search(str(formula))
            if match:
                hyperlinks_found += 1
                full_path = match.group(1)
                
                # Extract filename and directory path
                # Path uses backslashes: "M10 - GENERAL\P00 - ...\0854-10.00.001.01.PDF"
                # Replace backslashes with forward slashes for consistency
                full_path_unix = full_path.replace('\\', '/')
                parts = full_path_unix.split('/')
                filename = parts[-1]
                relative_dir_unix = '/'.join(parts[:-1])
                
                # Get other columns
                supergrandparent = sheet.cell(row=row, column=1).value  # Column A
                superparent = sheet.cell(row=row, column=2).value       # Column B
                revision = sheet.cell(row=row, column=4).value          # Column D
                doc_description = sheet.cell(row=row, column=5).value   # Column E
                supplier_code = sheet.cell(row=row, column=6).value     # Column F
                supplier_name = sheet.cell(row=row, column=7).value     # Column G
                status = sheet.cell(row=row, column=8).value            # Column H
                
                # Use lowercase filename as key for case-insensitive matching
                key = filename.lower()
                
                metadata[key] = {
                    'filename': filename,
                    'relative_path': full_path_unix,  # Use forward slashes
                    'relative_path_unix': full_path_unix,
                    'relative_dir': relative_dir_unix,
                    'relative_dir_unix': relative_dir_unix,
                    'document_description': doc_description,
                    'supergrandparent': supergrandparent,
                    'superparent': superparent,
                    'revision': str(revision) if revision else None,
                    'supplier_code': supplier_code,
                    'supplier_name': supplier_name,
                    'status': status,
                }
    
    wb.close()
    
    print(f"Processed {rows_processed} rows")
    print(f"Found {hyperlinks_found} HYPERLINK formulas")
    print(f"Loaded metadata for {len(metadata)} unique files")
    
    return metadata


def save_metadata_cache(metadata: dict, cache_path: str = "drawing_metadata.pkl"):
    """Save metadata to a pickle file for fast loading."""
    with open(cache_path, 'wb') as f:
        pickle.dump(metadata, f)
    print(f"Saved metadata cache to {cache_path}")


def load_metadata_cache(cache_path: str = "drawing_metadata.pkl") -> dict:
    """Load metadata from pickle cache."""
    if os.path.exists(cache_path):
        with open(cache_path, 'rb') as f:
            return pickle.load(f)
    return {}


def main():
    excel_path = sys.argv[1] if len(sys.argv) > 1 else "Overview_Drawings_and_Documents_0854_-_Boreas.xlsx"
    output_path = sys.argv[2] if len(sys.argv) > 2 else "drawing_metadata.pkl"
    
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found: {excel_path}")
        sys.exit(1)
    
    print("=" * 60)
    print("Ship Cable Database - Import Drawing Metadata")
    print("=" * 60)
    
    # Import metadata
    metadata = import_drawing_metadata(excel_path)
    
    # Save to cache for use by indexer
    save_metadata_cache(metadata, output_path)
    
    print("\n" + "=" * 60)
    print("Import Complete!")
    print("=" * 60)
    print(f"Documents with metadata: {len(metadata)}")
    
    # Show some examples
    print("\nSample entries:")
    for i, (filename, info) in enumerate(list(metadata.items())[:5]):
        print(f"\n  Filename: {info['filename']}")
        print(f"  Description: {info['document_description']}")
        print(f"  Path: {info['relative_dir_unix']}")
    
    # Show electrical drawings specifically
    electrical = {k: v for k, v in metadata.items() if 'M40' in str(v.get('supergrandparent', ''))}
    print(f"\n\nElectrical drawings (M40): {len(electrical)}")
    for i, (filename, info) in enumerate(list(electrical.items())[:3]):
        print(f"  {info['filename']}: {info['document_description']}")


if __name__ == "__main__":
    main()
