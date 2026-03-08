import re
import sqlite3

import click
import pandas as pd
from rich.console import Console
from rich.table import Table

from setup.import_equipment_and_cables import (
    _build_col_map,
    _print_columns,
    _prompt_column,
    _val,
)

DEFAULT_HYPERLINK_RE = r'=HYPERLINK\("([^"]+)"'


def _preview_table(records: list[dict]) -> None:
    if not records:
        click.echo("  (no records to preview)")
        return
    table = Table(show_header=True, header_style="bold", show_lines=True)
    for col in records[0]:
        table.add_column(col, overflow="fold")
    for record in records[:5]:
        table.add_row(*[str(v) if v is not None else "" for v in record.values()])
    Console().print(table)


def _extract_path(cell_value: object, pattern: re.Pattern) -> str | None:
    """Extract and normalise a path using the given compiled regex, or use the
    raw value as-is if the pattern does not match."""
    if not isinstance(cell_value, str):
        return None
    s = cell_value.strip()
    if not s:
        return None
    m = pattern.search(s)
    path = m.group(1) if m else s
    return path.replace("\\", "/")


def import_metadata(conn: sqlite3.Connection, filepath: str) -> None:
    """Import drawing metadata from an Excel or CSV file and update matching documents."""

    # ── Step 1: load file ────────────────────────────────────────────────────
    filepath = filepath.replace("\\", "/")
    ext = filepath.rsplit(".", 1)[-1].lower()

    header_row = click.prompt(
        "  Header row (1 = first row is header, 4 = fourth row is header, 0 = no header)",
        default=4,
        type=int,
    )
    pandas_header = None if header_row == 0 else header_row - 1

    try:
        if ext in ("xlsx", "xls"):
            # Use openpyxl with data_only=False so that HYPERLINK formula strings
            # are preserved as text rather than replaced with cached display values.
            from openpyxl import load_workbook
            wb = load_workbook(filepath, data_only=False)
            ws = wb.active
            all_rows = [
                [cell.value for cell in row]
                for row in ws.iter_rows()
            ]
            if pandas_header is None:
                df = pd.DataFrame(all_rows)
            else:
                df = pd.DataFrame(
                    all_rows[pandas_header + 1:],
                    columns=all_rows[pandas_header],
                )
        elif ext == "csv":
            df = pd.read_csv(filepath, header=pandas_header)
        else:
            click.echo(f"Unsupported file type: .{ext}  (expected .xlsx, .xls, or .csv)")
            return
    except FileNotFoundError:
        click.echo(f"File not found: {filepath}")
        return
    except Exception as e:
        click.echo(f"Failed to load file: {e}")
        return

    click.echo(f"  Loaded {len(df)} rows.")

    # ── Step 2: column mapping ───────────────────────────────────────────────
    _print_columns(df)
    col_map = _build_col_map(df)

    click.echo("\nMap fields to columns (Enter to accept default, letter or name to override):")

    path_col          = _prompt_column("relative_path        [required]", "Filename",               col_map, df, required=True)
    desc_col          = _prompt_column("document_description",             "Document Description",   col_map, df, required=False)
    supplier_code_col = _prompt_column("supplier_code",                    "Supplier Document Code", col_map, df, required=False)
    supplier_name_col = _prompt_column("supplier_name",                    "Supplier Name",          col_map, df, required=False)
    sgp_col           = _prompt_column("supergrandparent",                 "Supergrandparent",       col_map, df, required=False)
    sp_col            = _prompt_column("superparent",                      "Superparent",            col_map, df, required=False)
    revision_col      = _prompt_column("revision",                         "Revision",               col_map, df, required=False)
    status_col        = _prompt_column("status",                           "Status",                 col_map, df, required=False)

    # ── Step 3: regex for path extraction ────────────────────────────────────
    if path_col is not None:
        sample = df[path_col].dropna().head(3).tolist()
        click.echo(f"\n  Sample path column values: {sample}")

    while True:
        raw_re = click.prompt(
            "\n  Regex to extract path (group 1 is used; Enter to use default, '-' to use raw value)",
            default=DEFAULT_HYPERLINK_RE,
        )
        if raw_re == "-":
            path_pattern = re.compile(r"(?s).*")  # dummy — group 1 won't match, raw value used
            use_raw = True
            break
        try:
            path_pattern = re.compile(raw_re, re.IGNORECASE)
            # Verify it has a capturing group
            if path_pattern.groups < 1:
                click.echo("  Pattern must contain at least one capturing group, e.g. ([^\"]+)")
                continue
            use_raw = False
            break
        except re.error as e:
            click.echo(f"  Invalid regex: {e}")

    def extract(cell_value: object) -> str | None:
        if use_raw:
            if not isinstance(cell_value, str):
                return None
            s = cell_value.strip()
            return s.replace("\\", "/") if s else None
        return _extract_path(cell_value, path_pattern)

    # ── Step 4: build records ─────────────────────────────────────────────────
    records: list[dict] = []
    for _, row in df.iterrows():
        raw_path = row.get(path_col) if path_col is not None else None
        relative_path = extract(raw_path)
        if not relative_path:
            continue
        records.append({
            "relative_path":        relative_path,
            "document_description": _val(row, desc_col),
            "supplier_code":        _val(row, supplier_code_col),
            "supplier_name":        _val(row, supplier_name_col),
            "supergrandparent":     _val(row, sgp_col),
            "superparent":          _val(row, sp_col),
            "revision":             _val(row, revision_col),
            "status":               _val(row, status_col),
        })

    # ── Step 5: preview ───────────────────────────────────────────────────────
    click.echo("\nFirst 5 rows of mapped data:")
    _preview_table(records)

    if not click.confirm("\nProceed with update?", default=True):
        click.echo("Update cancelled.")
        return

    # ── Step 6: update in single transaction ─────────────────────────────────
    matched = 0
    unmatched = 0

    DB_FIELDS = [
        "document_description",
        "supplier_code",
        "supplier_name",
        "supergrandparent",
        "superparent",
        "revision",
        "status",
    ]

    with conn:
        for record in records:
            updates = {k: record[k] for k in DB_FIELDS if record[k] is not None}
            if not updates:
                continue
            set_clause = ", ".join(f"{col} = ?" for col in updates)
            params = list(updates.values()) + [record["relative_path"]]
            cur = conn.execute(
                f"UPDATE documents SET {set_clause} WHERE relative_path = ?",
                params,
            )
            if cur.rowcount > 0:
                matched += 1
            else:
                unmatched += 1

    # ── Step 7: summary ───────────────────────────────────────────────────────
    click.echo(f"\nUpdate complete: {matched} matched and updated  |  {unmatched} had no match")
